"""Genera el Excel de resultados: 3 hojas — Decisores / Revisar / Sin resultado.

- Columna **Teléfono** vacía (la rellena el Paso 4 cuando exista).
- LinkedIn (persona y empresa) como **hipervínculos**.
- Clasificación con color: **verde** = decisor, **ámbar** = revisar.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Classification, Company, CompanyStatus, Contact
from .resolver.base import handle_of

# ---------------------------------------------------------------- estilos
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100", bold=True)
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
AMBER_FONT = Font(color="9C6500", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")

CONTACT_HEADERS = [
    "CIF", "Razón social", "Empresa", "Nombre", "Apellidos", "Cargo",
    "Ubicación", "Teléfono", "LinkedIn (persona)", "LinkedIn Empresa",
    "Clasificación", "Verificación",
]
CONTACT_WIDTHS = [14, 28, 24, 16, 18, 32, 26, 16, 44, 38, 14, 16]
EMPTY_HEADERS = ["CIF", "Razón social", "Empresa", "LinkedIn Empresa", "Nota"]
EMPTY_WIDTHS = [14, 28, 26, 44, 80]

_DEFAULT_NOTE = (
    "0 resultados con sede en España bajo los filtros. Posibles causas: equipo "
    "centralizado fuera de España, empresa sin el rol buscado, o URL/handle que no "
    "apunta a la entidad española. Revisar manualmente o ampliar la búsqueda."
)


def _prettify(handle: str) -> str:
    h = re.sub(r"[-_]+", " ", handle).strip(" -")
    return h.title() if h else handle


def _company_display(company: Company | None) -> str:
    if company is None:
        return ""
    if company.razon_social:
        return company.razon_social
    h = handle_of(company.linkedin_url)
    return _prettify(h) if h else company.raw_input


def _set_link(cell, url: str | None) -> None:
    if url:
        cell.value = url
        cell.hyperlink = url
        cell.font = LINK_FONT


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def company_key(company: Company) -> str:
    """Id estable de una empresa: su id de BD, o un id local derivado del handle."""
    return company.id or f"local:{handle_of(company.linkedin_url) or company.raw_input}"


def _write_contact_sheet(ws, contacts: list[Contact], by_id: dict[str, Company]) -> None:
    _header(ws, CONTACT_HEADERS)
    for c in contacts:
        company = by_id.get(c.company_id or "")
        ws.append([
            (company.cif if company else "") or "",
            (company.razon_social if company else "") or "",
            _company_display(company),
            c.first_name or "",
            c.last_name or "",
            c.title or "",
            c.location or "",
            "",  # Teléfono — vacío (Paso 4)
            "",  # LinkedIn persona (hipervínculo abajo)
            "",  # LinkedIn empresa (hipervínculo abajo)
            "Decisor" if c.classification == Classification.decisor else "Revisar",
            c.verify_flag or "",
        ])
        row = ws.max_row
        _set_link(ws.cell(row=row, column=9), c.linkedin_url)
        _set_link(
            ws.cell(row=row, column=10),
            c.company_linkedin_url or (company.linkedin_url if company else None),
        )
        cls = ws.cell(row=row, column=11)
        if c.classification == Classification.decisor:
            cls.fill, cls.font = GREEN_FILL, GREEN_FONT
        else:
            cls.fill, cls.font = AMBER_FILL, AMBER_FONT
    _autofit(ws, CONTACT_WIDTHS)


def export_excel(
    companies: list[Company],
    contacts: list[Contact],
    output_path: str | Path,
    *,
    area_name: str | None = None,
) -> Path:
    """Escribe el .xlsx de 3 hojas y devuelve la ruta. No descarta a nadie.

    Cada `contact.company_id` debe coincidir con `company_key(company)` (el pipeline
    lo garantiza, tanto en modo BD como en modo CLI sin BD).
    """
    by_id: dict[str, Company] = {company_key(c): c for c in companies}

    decisores = [c for c in contacts if c.classification == Classification.decisor]
    revisar = [c for c in contacts if c.classification == Classification.revisar]

    wb = Workbook()
    ws_dec = wb.active
    ws_dec.title = "Decisores"
    _write_contact_sheet(ws_dec, decisores, by_id)

    ws_rev = wb.create_sheet("Revisar")
    _write_contact_sheet(ws_rev, revisar, by_id)

    # Hoja "Sin resultado": empresas sin contactos o sin URL
    with_contacts = {c.company_id for c in contacts}
    ws_no = wb.create_sheet("Sin resultado")
    _header(ws_no, EMPTY_HEADERS)
    for comp in companies:
        empty = comp.status in (CompanyStatus.no_result, CompanyStatus.no_url, CompanyStatus.error)
        key = company_key(comp)
        if not empty and key in with_contacts:
            continue
        ws_no.append([
            comp.cif or "",
            comp.razon_social or "",
            _company_display(comp),
            "",
            comp.note or _DEFAULT_NOTE,
        ])
        _set_link(ws_no.cell(row=ws_no.max_row, column=4), comp.linkedin_url)
        ws_no.cell(row=ws_no.max_row, column=5).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
    _autofit(ws_no, EMPTY_WIDTHS)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
